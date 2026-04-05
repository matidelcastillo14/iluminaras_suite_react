from __future__ import annotations

import os
import re

from datetime import timezone as dt_timezone

from flask import Blueprint, render_template, jsonify, current_app, request, send_file
from flask_login import login_required

from ...models import BatchOrder, ImportedBatch, Setting
from ...services.batch_poller import poll_once, ensure_min_rows
from ...services.timezone import get_app_timezone_name, get_app_timezone
from ...utils import view_required

bp = Blueprint("batch_pedidos", __name__, url_prefix="/inventario/batch-pedidos")


def _iso_utc(dt):
    """Devuelve ISO 8601 en UTC con sufijo 'Z'.

    En la DB se guarda naive en UTC (datetime.utcnow). En caso de venir naive, asumimos UTC.
    """
    if not dt:
        return None
    if getattr(dt, "tzinfo", None) is None:
        dt = dt.replace(tzinfo=dt_timezone.utc)
    else:
        dt = dt.astimezone(dt_timezone.utc)
    return dt.isoformat().replace("+00:00", "Z")


@bp.get("/")
@login_required
@view_required("batch_pedidos")
def index():
    return render_template("batch_pedidos/index.html", app_timezone=get_app_timezone_name())


def _get_default_page_size() -> int:
    """Lee el page_size default desde Admin > Configuración o config."""
    try:
        row = Setting.query.filter_by(key="BATCH_PEDIDOS_ROWS_LIMIT").first()
        raw = (row.value if row else None)
        ps = int(raw if raw not in (None, "") else current_app.config.get("BATCH_PEDIDOS_ROWS_LIMIT", 80) or 80)
    except Exception:
        ps = 80
    return max(10, min(ps, 5000))


def _parse_list_params():
    """Soporta GET (querystring) o POST (json)."""
    payload = request.get_json(silent=True) if request.method == "POST" else None
    if not isinstance(payload, dict):
        payload = {}

    def _get(name: str, default=None):
        if name in payload:
            return payload.get(name, default)
        return request.args.get(name, default)

    try:
        page = int(_get("page", 1) or 1)
    except Exception:
        page = 1
    try:
        page_size = int(_get("page_size", _get_default_page_size()) or _get_default_page_size())
    except Exception:
        page_size = _get_default_page_size()

    page = max(1, page)
    page_size = max(10, min(page_size, 5000))

    sort_key = (_get("sort_key", "hora") or "hora").strip()
    sort_dir = (_get("sort_dir", "desc") or "desc").strip().lower()
    if sort_dir not in ("asc", "desc"):
        sort_dir = "desc"

    filters = _get("filters", {})
    if isinstance(filters, str):
        # intentar JSON en querystring
        try:
            import json

            filters = json.loads(filters)
        except Exception:
            filters = {}
    if not isinstance(filters, dict):
        filters = {}

    # Normalizar: mantenemos compatibilidad con el formato anterior:
    #   key -> list[str]
    # y soportamos formato "Excel":
    #   key -> {mode: 'values'|'text'|'range'|'date_range'|'time_range', ...}
    norm_filters: dict[str, object] = {}
    for k, v in filters.items():
        kk = str(k)
        if v is None:
            continue
        if isinstance(v, dict):
            norm_filters[kk] = v
            continue
        if isinstance(v, (list, tuple)):
            norm_filters[kk] = [str(x) for x in v if x is not None]
        else:
            norm_filters[kk] = [str(v)]

    return page, page_size, sort_key, sort_dir, norm_filters


def _apply_text_filter(q, col, spec: dict):
    """Filtro estilo Excel: contiene/empieza/igual/no contiene."""
    op = (spec.get("op") or "contains").strip().lower()
    val = (spec.get("value") or "").strip()
    if not val:
        return q
    if op == "equals":
        return q.filter(col == val)
    if op == "startswith":
        return q.filter(col.ilike(f"{val}%"))
    if op == "not_contains":
        return q.filter(~col.ilike(f"%{val}%"))
    # default: contains
    return q.filter(col.ilike(f"%{val}%"))


def _apply_numeric_range(q, col, spec: dict):
    lo = spec.get("from")
    hi = spec.get("to")
    try:
        lo_n = float(lo) if lo not in (None, "") else None
    except Exception:
        lo_n = None
    try:
        hi_n = float(hi) if hi not in (None, "") else None
    except Exception:
        hi_n = None
    if lo_n is not None:
        q = q.filter(col >= lo_n)
    if hi_n is not None:
        q = q.filter(col <= hi_n)
    return q


def _apply_local_dt_range(q, col, spec: dict, kind: str):
    """Filtra un datetime UTC-naive almacenado en DB por rango local.

    kind:
      - 'date' -> spec{from: 'YYYY-MM-DD', to:'YYYY-MM-DD'}
      - 'time' -> spec{from: 'HH:MM', to:'HH:MM'} (sobre el día actual no tiene sentido solo, se aplica como
                 condición sobre el tiempo local, usando extract(hour/minute).)

    Para fecha usamos límites locales convertidos a UTC.
    Para hora usamos EXTRACT para mantener simple (y razonablemente rápido para volúmenes moderados).
    """
    tz = get_app_timezone()

    if kind == "date":
        from_s = (spec.get("from") or "").strip()
        to_s = (spec.get("to") or "").strip()
        if not from_s and not to_s:
            return q
        from_dt_utc = None
        to_dt_utc = None
        from datetime import datetime, time, timezone as dt_tz

        try:
            if from_s:
                y, m, d = [int(x) for x in from_s.split("-")]
                local_start = datetime(y, m, d, 0, 0, 0, tzinfo=tz)
                from_dt_utc = local_start.astimezone(dt_tz.utc).replace(tzinfo=None)
        except Exception:
            from_dt_utc = None

        try:
            if to_s:
                y, m, d = [int(x) for x in to_s.split("-")]
                # inclusive end: hasta fin del día
                local_end = datetime(y, m, d, 23, 59, 59, 999999, tzinfo=tz)
                to_dt_utc = local_end.astimezone(dt_tz.utc).replace(tzinfo=None)
        except Exception:
            to_dt_utc = None

        if from_dt_utc is not None:
            q = q.filter(col >= from_dt_utc)
        if to_dt_utc is not None:
            q = q.filter(col <= to_dt_utc)
        return q

    if kind == "time":
        # Hora local independiente del día (ej: 08:00 a 12:30)
        from_s = (spec.get("from") or "").strip()
        to_s = (spec.get("to") or "").strip()
        if not from_s and not to_s:
            return q

        # Usamos SQL expression AT TIME ZONE para extraer hora/min local
        # Nota: esto requiere Postgres (Suite2 ya usa Postgres).
        from sqlalchemy import func, text

        local_ts = func.timezone(get_app_timezone_name(), col)
        # local_ts queda timestamp with time zone en el TZ elegido, extraemos hora/min
        def _parse_hm(s: str):
            try:
                hh, mm = s.split(":")
                return int(hh), int(mm)
            except Exception:
                return None

        lo = _parse_hm(from_s) if from_s else None
        hi = _parse_hm(to_s) if to_s else None

        if lo:
            hh, mm = lo
            q = q.filter(func.extract("hour", local_ts) * 60 + func.extract("minute", local_ts) >= (hh * 60 + mm))
        if hi:
            hh, mm = hi
            q = q.filter(func.extract("hour", local_ts) * 60 + func.extract("minute", local_ts) <= (hh * 60 + mm))
        return q

    return q


def _build_query(filters: dict[str, object], sort_key: str, sort_dir: str):
    q = BatchOrder.query.join(ImportedBatch, BatchOrder.imported_batch_id == ImportedBatch.id)

    # Normalización SQL del número de factura para que el filtrado coincida con lo que se muestra en UI
    # (ej: 'e-TK A0320206' -> 'A320206').
    # Regla: tomar la primera letra A-Z que esté seguida por dígitos, eliminar ceros inmediatamente
    # después de la letra, y conservar el resto de dígitos.
    from sqlalchemy import func

    def _invoice_norm_expr(col):
        src = func.upper(func.coalesce(col, ""))
        # Si no matchea, devuelve el string original. Para nuestros datos, esto es suficiente.
        return func.regexp_replace(src, r".*?([A-Z])0*([0-9]+).*", r"\\1\\2")

    def _as_list(v: object) -> list[str]:
        if v is None:
            return []
        if isinstance(v, (list, tuple)):
            return [str(x) for x in v if x is not None]
        return [str(v)]

    def _apply_text(col, f: dict):
        op = (f.get("op") or "contains").lower()
        val = str(f.get("value") or "").strip()
        if not val:
            return None
        if op == "eq" or op == "=":
            return col == val
        if op == "starts" or op == "startswith":
            return col.ilike(f"{val}%")
        if op in ("not_contains", "notcontains"):
            return ~col.ilike(f"%{val}%")
        # default contains
        return col.ilike(f"%{val}%")

    def _apply_numeric_range(col, f: dict):
        vmin = f.get("min")
        vmax = f.get("max")
        conds = []
        try:
            if vmin not in (None, ""):
                conds.append(col >= float(vmin))
        except Exception:
            pass
        try:
            if vmax not in (None, ""):
                conds.append(col <= float(vmax))
        except Exception:
            pass
        if not conds:
            return None
        from sqlalchemy import and_

        return and_(*conds)

    def _apply_local_dt_range(col, f: dict):
        """Filtro por rango de fecha/hora en timezone de la app.

        Espera:
          {from_date:'YYYY-MM-DD', to_date:'YYYY-MM-DD', from_time:'HH:MM', to_time:'HH:MM'}
        Cualquier combinación es válida.
        """
        from datetime import datetime, timezone as dt_tz
        from sqlalchemy import and_

        tz = get_app_timezone()
        fd = str(f.get("from_date") or "").strip()
        td = str(f.get("to_date") or "").strip()
        ft = str(f.get("from_time") or "").strip()
        tt = str(f.get("to_time") or "").strip()

        def _mk(d: str, t: str, end: bool) -> datetime | None:
            if not d:
                return None
            if not t:
                t = "23:59:59" if end else "00:00:00"
            else:
                # HH:MM -> HH:MM:SS
                t = t + ":00" if len(t) == 5 else t
            try:
                loc = datetime.fromisoformat(f"{d}T{t}")
            except Exception:
                return None
            loc = loc.replace(tzinfo=tz)
            return loc.astimezone(dt_tz.utc).replace(tzinfo=None)

        dt_from = _mk(fd, ft, end=False)
        dt_to = _mk(td, tt, end=True)
        conds = []
        if dt_from is not None:
            conds.append(col >= dt_from)
        if dt_to is not None:
            conds.append(col <= dt_to)
        return and_(*conds) if conds else None

    # filtros (server-side)
    for key, raw_filter in (filters or {}).items():
        if raw_filter is None:
            continue

        # Formato anterior: lista de valores exactos
        if not isinstance(raw_filter, dict):
            vals = _as_list(raw_filter)
            if not vals:
                continue
            if key == "batch_name":
                q = q.filter(ImportedBatch.batch_name.in_(vals))
            elif key == "id_web":
                q = q.filter(BatchOrder.id_web.in_(vals))
            elif key == "id_melicart":
                q = q.filter(BatchOrder.id_melicart.in_(vals))
            elif key == "id_meli":
                q = q.filter(BatchOrder.id_meli.in_(vals))
            elif key == "cliente":
                q = q.filter(BatchOrder.cliente.in_(vals))
            elif key == "estado_factura":
                q = q.filter(BatchOrder.estado_factura.in_(vals))
            elif key == "n_factura":
                inv = _invoice_norm_expr(BatchOrder.n_factura)
                empty_selected = "" in vals
                non_empty = [v.upper() for v in vals if v not in (None, "")]
                if non_empty and empty_selected:
                    from sqlalchemy import or_

                    q = q.filter(or_(inv.in_(non_empty), BatchOrder.n_factura.is_(None), BatchOrder.n_factura == ""))
                elif non_empty:
                    q = q.filter(inv.in_(non_empty))
                elif empty_selected:
                    from sqlalchemy import or_

                    q = q.filter(or_(BatchOrder.n_factura.is_(None), BatchOrder.n_factura == ""))
            elif key == "monto_compra":
                nums = []
                for x in vals:
                    try:
                        nums.append(float(x))
                    except Exception:
                        pass
                if nums:
                    q = q.filter(BatchOrder.monto_compra.in_(nums))
            continue

        # Formato Excel (dict)
        mode = (raw_filter.get("mode") or "values").lower()

        if mode == "values":
            vals = _as_list(raw_filter.get("values"))
            if not vals:
                continue
            if key == "batch_name":
                q = q.filter(ImportedBatch.batch_name.in_(vals))
            elif key == "id_web":
                q = q.filter(BatchOrder.id_web.in_(vals))
            elif key == "id_melicart":
                q = q.filter(BatchOrder.id_melicart.in_(vals))
            elif key == "id_meli":
                q = q.filter(BatchOrder.id_meli.in_(vals))
            elif key == "cliente":
                q = q.filter(BatchOrder.cliente.in_(vals))
            elif key == "estado_factura":
                q = q.filter(BatchOrder.estado_factura.in_(vals))
            elif key == "n_factura":
                norm = _invoice_norm_expr(BatchOrder.n_factura)
                # soportar '(Vacío)' => valor '' desde UI
                want_empty = "" in vals
                non_empty = [v.upper() for v in vals if v not in (None, "")]
                if non_empty and want_empty:
                    from sqlalchemy import or_

                    q = q.filter(or_(norm.in_(non_empty), BatchOrder.n_factura.is_(None), BatchOrder.n_factura == ""))
                elif non_empty:
                    q = q.filter(norm.in_(non_empty))
                elif want_empty:
                    from sqlalchemy import or_

                    q = q.filter(or_(BatchOrder.n_factura.is_(None), BatchOrder.n_factura == ""))
            continue

        if mode == "text":
            expr = None
            if key == "batch_name":
                expr = _apply_text(ImportedBatch.batch_name, raw_filter)
            elif key == "id_web":
                expr = _apply_text(BatchOrder.id_web, raw_filter)
            elif key == "id_melicart":
                expr = _apply_text(BatchOrder.id_melicart, raw_filter)
            elif key == "id_meli":
                expr = _apply_text(BatchOrder.id_meli, raw_filter)
            elif key == "cliente":
                expr = _apply_text(BatchOrder.cliente, raw_filter)
            elif key == "estado_factura":
                expr = _apply_text(BatchOrder.estado_factura, raw_filter)
            elif key == "n_factura":
                # Filtrar contra la versión normalizada (misma que se muestra en UI)
                op = (raw_filter.get("op") or "contains").lower()
                val = str(raw_filter.get("value") or "").strip().upper()
                if val:
                    norm = _invoice_norm_expr(BatchOrder.n_factura)
                    if op in ("eq", "="):
                        expr = norm == val
                    elif op in ("starts", "startswith"):
                        expr = norm.like(f"{val}%")
                    elif op in ("not_contains", "notcontains"):
                        expr = ~norm.like(f"%{val}%")
                    else:
                        expr = norm.like(f"%{val}%")
            if expr is not None:
                q = q.filter(expr)
            continue

        if mode in ("range", "numeric_range") and key == "monto_compra":
            expr = _apply_numeric_range(BatchOrder.monto_compra, raw_filter)
            if expr is not None:
                q = q.filter(expr)
            continue

        if mode in ("datetime_range", "dt_range") and key in ("hora", "order_date"):
            col = BatchOrder.created_at if key == "hora" else BatchOrder.order_date
            expr = _apply_local_dt_range(col, raw_filter)
            if expr is not None:
                q = q.filter(expr)
            continue

    # sort (server-side)
    sort_map = {
        "hora": BatchOrder.created_at,
        "order_date": BatchOrder.order_date,
        "batch_name": ImportedBatch.batch_name,
        "id_web": BatchOrder.id_web,
        "id_melicart": BatchOrder.id_melicart,
        "id_meli": BatchOrder.id_meli,
        "cliente": BatchOrder.cliente,
        "monto_compra": BatchOrder.monto_compra,
        # Ordenar por lo que se muestra al usuario
        "n_factura": _invoice_norm_expr(BatchOrder.n_factura),
    }
    col = sort_map.get(sort_key, BatchOrder.created_at)
    q = q.order_by(col.asc() if sort_dir == "asc" else col.desc(), BatchOrder.id.desc())

    return q


@bp.route("/api/list", methods=["GET", "POST"])
@login_required
@view_required("batch_pedidos")
def api_list():
    page, page_size, sort_key, sort_dir, filters = _parse_list_params()

    # backfill acotado solo para asegurar que haya datos (especialmente en instalaciones nuevas)
    try:
        if BatchOrder.query.count() < (page_size * 2):
            ensure_min_rows(page_size * 2, max_batches=5)
    except Exception:
        pass

    q = _build_query(filters, sort_key, sort_dir)
    total = q.count()
    rows = q.offset((page - 1) * page_size).limit(page_size).all()

    tz = get_app_timezone()

    items = []
    for r in rows:
        # Local (para UI/inputs date/time)
        created_loc = None
        if r.created_at:
            try:
                created_loc = (r.created_at.replace(tzinfo=dt_timezone.utc)).astimezone(tz)
            except Exception:
                created_loc = None

        order_dt = getattr(r, "order_date", None)
        order_loc = None
        if order_dt:
            try:
                order_loc = (order_dt.replace(tzinfo=dt_timezone.utc)).astimezone(tz)
            except Exception:
                order_loc = None

        items.append(
            {
                "hora": _iso_utc(r.created_at),
                "fecha": created_loc.date().isoformat() if created_loc else "",
                "hora_solo": created_loc.strftime("%H:%M") if created_loc else "",
                "order_date": _iso_utc(order_dt),
                "order_fecha": order_loc.date().isoformat() if order_loc else "",
                "order_hora": order_loc.strftime("%H:%M") if order_loc else "",
                "batch_name": r.batch.batch_name if r.batch else None,
                "batch_odoo_id": r.batch.odoo_batch_id if r.batch else None,
                "id_web": r.id_web or "",
                "id_melicart": r.id_melicart or "",
                "id_meli": r.id_meli or "",
                "cliente": r.cliente or "",
                "monto_compra": float(r.monto_compra) if r.monto_compra is not None else None,
                "n_factura_raw": r.n_factura or "",
                "n_factura_fmt": (_normalize_invoice_code(r.n_factura or "") or (r.n_factura or "")),
                "estado_factura": r.estado_factura or "",
                "link_factura": r.link_factura or "",
                "sale_order_id": r.sale_order_id,
                "sale_order_ref": r.sale_order_ref or "",
            }
        )

    return jsonify(
        {
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size,
            "sort_key": sort_key,
            "sort_dir": sort_dir,
        }
    )




@bp.get("/api/filter_options")
@login_required
@view_required("batch_pedidos")
def api_filter_options():
    """
    Devuelve opciones de filtros (server-side), para que el dropdown no dependa de la página actual.
    Params:
      field: por ahora soporta "batch_name"
      q: texto de búsqueda (opcional)
      limit: máximo de resultados (default 500, max 2000)
    """
    field = (request.args.get("field") or "").strip()
    q = (request.args.get("q") or "").strip()
    try:
        limit = int(request.args.get("limit", 500) or 500)
    except Exception:
        limit = 500
    limit = max(10, min(limit, 2000))

    if field != "batch_name":
        return jsonify({"error": "field no soportado"}), 400

    qry = ImportedBatch.query
    if q:
        # búsqueda case-insensitive
        qry = qry.filter(ImportedBatch.batch_name.ilike(f"%{q}%"))
    # distinct + orden estable
    qry = qry.with_entities(ImportedBatch.batch_name).distinct().order_by(ImportedBatch.batch_name.desc()).limit(limit)

    vals = [r[0] for r in qry.all() if r and r[0] is not None]
    return jsonify({"field": field, "items": vals})


@bp.post("/api/export_xlsx")
@login_required
@view_required("batch_pedidos")
def api_export_xlsx():
    """Exporta a XLSX usando filtros/orden actuales (sobre TODO el set, no solo la página)."""
    payload = request.get_json(silent=True) or {}
    if not isinstance(payload, dict):
        payload = {}
    filters = payload.get("filters") or {}
    sort_key = (payload.get("sort_key") or "hora").strip()
    sort_dir = (payload.get("sort_dir") or "desc").strip().lower()
    if sort_dir not in ("asc", "desc"):
        sort_dir = "desc"
    if not isinstance(filters, dict):
        filters = {}

    q = _build_query(filters, sort_key, sort_dir)
    # hard limit para evitar abusos
    rows_db = q.limit(20000).all()
    if not rows_db:
        return jsonify({"error": "No hay filas para exportar"}), 404

    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos"

    headers = [
        "Fecha y hora",
        "Fecha pedido (Odoo)",
        "ID Batch",
        "ID Web",
        "ID MeliCart",
        "ID Meli",
        "Cliente",
        "Monto",
        "N° Factura",
        "Ver",
    ]
    ws.append(headers)

    for r in rows_db:
        ws.append(
            [
                _iso_utc(r.created_at) or "",
                _iso_utc(getattr(r, "order_date", None)) or "",
                (r.batch.batch_name if r.batch else "") or "",
                r.id_web or "",
                r.id_melicart or "",
                r.id_meli or "",
                r.cliente or "",
                float(r.monto_compra) if r.monto_compra is not None else None,
                (_normalize_invoice_code(r.n_factura or "") or (r.n_factura or "")),
                r.link_factura or "",
            ]
        )

    # autosize simple
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for cell in ws[get_column_letter(col_idx)]:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 60)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    from datetime import datetime
    filename = f"pedidos_por_batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )





def _normalize_invoice_code(raw: str) -> str | None:
    '''Normaliza un número de factura del batch al código usado en el PDF.

    Ejemplos:
    - 'e-TK A0320206' -> 'A320206'
    - 'A0320120' -> 'A320120'

    Regla: tomar desde la letra (A-Z) y eliminar los ceros inmediatamente después
    de esa letra, hasta el primer dígito > 0. Luego asegurar formato letra + 6 dígitos
    (como los últimos 7 chars del nombre del PDF). Si no se puede, devuelve None.
    '''
    if not raw:
        return None
    s = str(raw).strip()

    # Buscar: primera letra + dígitos (con o sin prefijos como 'e-TK ')
    m = re.search(r"([A-Za-z])\s*([0-9]{1,12})", s)
    if not m:
        return None
    letter = m.group(1).upper()
    digits = m.group(2)

    # Eliminar ceros inmediatamente después de la letra
    digits = digits.lstrip("0")
    if not digits:
        return None

    # Para el PDF se usa letra + 6 dígitos. Si viene más largo, tomar los últimos 6.
    if len(digits) > 6:
        digits = digits[-6:]

    if len(digits) != 6 or not digits.isdigit():
        return None

    return f"{letter}{digits}"


@bp.post("/api/export_facturas_zip")
@login_required
@view_required("batch_pedidos")
def api_export_facturas_zip():
    """Exporta facturas PDF en un ZIP usando filtros/orden actuales (sobre TODO el set)."""
    payload = request.get_json(silent=True) or {}
    if not isinstance(payload, dict):
        payload = {}
    filters = payload.get("filters") or {}
    sort_key = (payload.get("sort_key") or "hora").strip()
    sort_dir = (payload.get("sort_dir") or "desc").strip().lower()
    if sort_dir not in ("asc", "desc"):
        sort_dir = "desc"
    if not isinstance(filters, dict):
        filters = {}

    q = _build_query(filters, sort_key, sort_dir)
    rows_db = q.limit(20000).all()
    if not rows_db:
        return jsonify({"error": "No hay filas para exportar"}), 404

    # Directorio de facturas (legacy)
    base_dir = os.path.abspath(os.path.join(current_app.root_path, os.pardir))
    gen_dir = os.path.join(base_dir, "legacy_apps", "cfe_auto", "generated")

    if not os.path.isdir(gen_dir):
        return jsonify({"error": f"No existe el directorio de facturas: {gen_dir}"}), 500

    # armar mapa code->filepath (solo factura, no CAMBIO)
    code_to_path: dict[str, str] = {}
    try:
        for name in os.listdir(gen_dir):
            if not name.lower().endswith(".pdf"):
                continue
            if "_cambio" in name.lower():
                continue
            stem = name[:-4]
            if len(stem) < 7:
                continue
            code = stem[-7:].upper()
            # validar forma letra+6 dígitos
            if not re.match(r"^[A-Z][0-9]{6}$", code):
                continue
            # elegir el primero (determinístico por orden alfabético)
            if code not in code_to_path:
                code_to_path[code] = os.path.join(gen_dir, name)
    except Exception as e:
        return jsonify({"error": f"No se pudo leer el directorio de facturas: {e}"}), 500

    wanted: list[str] = []
    for r in rows_db:
        code = _normalize_invoice_code(r.n_factura or "")
        if code:
            wanted.append(code)

    # únicos, preservando orden
    seen = set()
    wanted_unique = []
    for c in wanted:
        if c not in seen:
            wanted_unique.append(c)
            seen.add(c)

    found_files = []
    missing = []
    for code in wanted_unique:
        p = code_to_path.get(code)
        if p and os.path.isfile(p):
            found_files.append((code, p))
        else:
            missing.append(code)

    if not found_files:
        return jsonify({"error": "No se encontraron PDFs para exportar", "missing": missing}), 404

    from io import BytesIO
    import zipfile
    from datetime import datetime

    bio = BytesIO()
    with zipfile.ZipFile(bio, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for code, p in found_files:
            # mantener nombre original del PDF
            arcname = os.path.basename(p)
            # evitar colisiones: si se repite, agregar sufijo
            if arcname in zf.namelist():
                base, ext = os.path.splitext(arcname)
                arcname = f"{base}_{code}{ext}"
            zf.write(p, arcname=arcname)

        # incluir reporte de faltantes si los hay
        if missing:
            report = "FALTANTES\n" + "\n".join(missing) + "\n"
            zf.writestr("faltantes.txt", report)

    bio.seek(0)
    filename = f"facturas_batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    return send_file(bio, as_attachment=True, download_name=filename, mimetype="application/zip")
@bp.post("/api/poll_now")
@login_required
@view_required("batch_pedidos")
def api_poll_now():
    return jsonify(poll_once())

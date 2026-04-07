from __future__ import annotations

import os
import json
from datetime import datetime, date, time, timezone
from pathlib import Path

from flask import Blueprint, render_template, request, redirect, url_for, abort, jsonify, send_file, current_app
from flask_login import login_required, current_user

from ...extensions import db
from ...models import TrackingShipment, TrackingEvent, TrackingScan, BatchOrder
from ...utils import view_required, can_view
from ...services.tracking_service import add_event, record_scan, ventas_override_status, ventas_reset_to_zero
from ...services import etiquetas_legacy_engine
from ...services.odoo_readonly import get_order_full
from ...services.timezone import get_app_timezone
from ...services.tracking_labels import label_status


bp = Blueprint("rastreo", __name__, url_prefix="/rastreo")


def _configure_engine():
    etiquetas_legacy_engine.configure(current_app.config, root_path=str(Path(current_app.root_path).parent), logger=current_app.logger)


def _find_shipment_by_any_code(raw_code: str) -> TrackingShipment | None:
    """Resuelve un escaneo a shipment por:
    1) tracking_code (QR viejo / rastreo)
    2) order_name (QR nuevo: "S...")
    3) id_web (fallback)
    Soporta QR que viene como URL /rastreo/go/<code>.
    """
    code = _parse_scanned_code(raw_code or '')
    code = (code or '').strip()
    if not code:
        return None

    sh = TrackingShipment.query.filter_by(tracking_code=code).first()
    if sh:
        return sh
    sh = TrackingShipment.query.filter_by(order_name=code).first()
    if sh:
        return sh
    return TrackingShipment.query.filter_by(id_web=code).first()


def _shipment_by_code(code: str) -> TrackingShipment:
    sh = _find_shipment_by_any_code(code)
    if not sh:
        abort(404)
    return sh


@bp.get("/go/<tracking_code>")
@login_required
def go(tracking_code: str):
    # Redirige a la mejor vista disponible según permisos del usuario
    if can_view("rastreo_cadeteria"):
        return redirect(url_for("rastreo.cadeteria_detail", tracking_code=tracking_code))
    if can_view("rastreo_deposito"):
        return redirect(url_for("rastreo.deposito_detail", tracking_code=tracking_code))
    if can_view("rastreo_ventas"):
        return redirect(url_for("rastreo.ventas_detail", tracking_code=tracking_code))
    abort(403)


@bp.get("/deposito")
@login_required
@view_required("rastreo_deposito")
def deposito():
    return render_template("rastreo/scan.html", title="Rastreo - Depósito", mode="deposito")


@bp.post("/deposito")
@login_required
@view_required("rastreo_deposito")
def deposito_post():
    code = (request.form.get("code") or "").strip()
    if not code:
        return redirect(url_for("rastreo.deposito"))
    return redirect(url_for("rastreo.deposito_detail", tracking_code=code))


@bp.get("/deposito/<tracking_code>")
@login_required
@view_required("rastreo_deposito")
def deposito_detail(tracking_code: str):
    sh = _shipment_by_code(tracking_code)
    _configure_engine()
    order_info = {}
    try:
        order_info = etiquetas_legacy_engine._odoo_order_detail(sh.odoo_order_id)
    except Exception:
        order_info = {"pedido": sh.order_name}
    # Incluir líneas del pedido (para dropdown de faltantes)
    try:
        full = get_order_full(sh.odoo_order_id)
        # full['lines'] tiene name/product/qty
        order_info["lines"] = full.get("lines") or []
    except Exception:
        order_info.setdefault("lines", [])
    # Registrar visualización en depósito para trazabilidad
    try:
        record_scan(sh, user_id=current_user.id, source='deposito_view')
    except Exception:
        # no bloquear en caso de error
        pass
    events = TrackingEvent.query.filter_by(shipment_id=sh.id).order_by(TrackingEvent.created_at.desc()).all()
    return render_template("rastreo/detail.html", title="Rastreo - Depósito", mode="deposito", shipment=sh, order=order_info, events=events)


@bp.get("/deposito/devoluciones")
@login_required
@view_required("rastreo_deposito")
def deposito_devoluciones():
    # Devoluciones solicitadas (último evento = RETURN_TO_DEPOT_REQUESTED)
    sub = (
        db.session.query(
            TrackingEvent.shipment_id,
            db.func.max(TrackingEvent.id).label("max_id"),
        )
        .group_by(TrackingEvent.shipment_id)
        .subquery()
    )
    latest = (
        db.session.query(TrackingEvent)
        .join(sub, db.and_(TrackingEvent.shipment_id == sub.c.shipment_id, TrackingEvent.id == sub.c.max_id))
        .filter(TrackingEvent.event_type == "RETURN_TO_DEPOT_REQUESTED")
        .order_by(TrackingEvent.created_at.desc(), TrackingEvent.id.desc())
        .limit(500)
        .all()
    )

    items = []
    for ev in latest:
        payload = {}
        try:
            payload = json.loads(ev.payload_json) if ev.payload_json else {}
        except Exception:
            payload = {}
        if int(payload.get("return_to_user_id") or 0) != int(current_user.id):
            continue
        sh = TrackingShipment.query.get(ev.shipment_id)
        if not sh:
            continue
        items.append({
            "shipment": sh,
            "event": ev,
        })

    return render_template(
        "rastreo/devoluciones_pendientes.html",
        title="Rastreo - Depósito - Devoluciones pendientes",
        items=items,
    )


@bp.post("/deposito/devoluciones/<int:shipment_id>/confirm")
@login_required
@view_required("rastreo_deposito")
def deposito_devolucion_confirm(shipment_id: int):
    sh = TrackingShipment.query.get_or_404(int(shipment_id))
    # Validación: último evento debe ser RETURN_TO_DEPOT_REQUESTED y asignado a este usuario
    last = TrackingEvent.query.filter_by(shipment_id=sh.id).order_by(TrackingEvent.created_at.desc(), TrackingEvent.id.desc()).first()
    if not last or last.event_type != "RETURN_TO_DEPOT_REQUESTED":
        abort(400, "not_pending_return")
    try:
        payload = json.loads(last.payload_json) if last.payload_json else {}
    except Exception:
        payload = {}
    if int(payload.get("return_to_user_id") or 0) != int(current_user.id) and current_user.role != "admin":
        abort(403)

    try:
        add_event(sh, event_type="RETURNED", note="Recepción confirmada en depósito", payload={"source":"deposito_confirm"}, created_by_user_id=current_user.id)
    except ValueError as ex:
        db.session.rollback()
        abort(400, str(ex))

    return redirect(url_for("rastreo.deposito_devoluciones"))



@bp.post("/deposito/<tracking_code>/event")
@login_required
@view_required("rastreo_deposito")
def deposito_event(tracking_code: str):
    sh = _shipment_by_code(tracking_code)
    event_type = (request.form.get("event_type") or "").strip()
    note = (request.form.get("note") or "").strip() or None
    payload = None
    if event_type == "STOCK_MISSING":
        # Permitir recibir múltiples productos faltantes. El formulario puede enviar
        # "missing_item" único o "missing_items" como lista.
        # Filtrar y limpiar valores vacíos.
        items = []
        # getlist devuelve lista de valores para keys repetidas o sufijo []
        for val in request.form.getlist("missing_items"):
            v = (val or "").strip()
            if v:
                items.append(v)
        # Fallback al campo único legacy
        if not items:
            single = (request.form.get("missing_item") or "").strip()
            if single:
                items.append(single)
        # Campo adicional opcional (manual)
        extra = (request.form.get("missing_item_extra") or "").strip()
        if extra:
            items.append(extra)
        if not items:
            abort(400)
        # Construir payload con lista completa
        payload = {"missing_items": items}

    # Depósito: la última acción permitida es READY_FOR_DISPATCH.
    if event_type not in {"PICKING_STARTED", "READY_FOR_DISPATCH", "STOCK_MISSING"}:
        abort(400)
    try:
        # Para STOCK_MISSING necesitamos atomicidad: si falla el guardado de items
        # NO debe quedar el shipment en STOCK_MISSING.
        if event_type == "STOCK_MISSING":
            ev = add_event(sh, event_type=event_type, note=note, payload=payload, created_by_user_id=current_user.id, commit=False)
            from ...models import TrackingStockMissingItem
            for prod_name in (payload or {}).get("missing_items", []):
                mi = TrackingStockMissingItem(
                    shipment_id=sh.id,
                    event_id=ev.id,
                    product_name_snapshot=str(prod_name),
                )
                db.session.add(mi)
            db.session.commit()
        else:
            add_event(sh, event_type=event_type, note=note, payload=payload, created_by_user_id=current_user.id)
    except ValueError as ex:
        db.session.rollback()
        abort(400, str(ex))
    except Exception:
        # Cualquier error debe dejar la transacción sin cambios
        db.session.rollback()
        raise
    return redirect(url_for("rastreo.deposito_detail", tracking_code=tracking_code))


@bp.get("/cadeteria")
@login_required
@view_required("rastreo_cadeteria")
def cadeteria():
    return render_template("rastreo/cadeteria_workflow.html", title="Rastreo - Cadetería", mode="cadeteria")


@bp.post("/cadeteria")
@login_required
@view_required("rastreo_cadeteria")
def cadeteria_post():
    code = (request.form.get("code") or "").strip()
    if not code:
        return redirect(url_for("rastreo.cadeteria"))
    return redirect(url_for("rastreo.cadeteria_detail", tracking_code=code))




def _parse_scanned_code(raw: str) -> str:
    code = (raw or '').strip()
    if not code:
        return ''
    # si el QR contiene URL /rastreo/go/<code>, extraer último segmento
    if '/rastreo/go/' in code:
        try:
            code = code.split('/rastreo/go/')[-1].split('?')[0].strip('/')
        except Exception:
            pass
    return code.strip()


@bp.post('/cadeteria/api/batch_clear')
@login_required
@view_required('rastreo_cadeteria')
def cadeteria_batch_clear():
    TrackingScan.query.filter_by(user_id=current_user.id, source='batch').delete(synchronize_session=False)
    # también limpiar "active" para evitar estados colgados
    TrackingScan.query.filter_by(user_id=current_user.id, source='active').delete(synchronize_session=False)
    db.session.commit()
    return jsonify({'ok': True})


@bp.post('/cadeteria/api/batch_add')
@login_required
@view_required('rastreo_cadeteria')
def cadeteria_batch_add():
    data = request.get_json(silent=True) or {}
    raw = (data.get('code') or '').strip()
    code = _parse_scanned_code(raw)
    if not code:
        return jsonify({'ok': False, 'error': 'code_required'}), 400

    sh = _find_shipment_by_any_code(code)
    if not sh:
        return jsonify({'ok': False, 'error': 'not_found'}), 404

    # evitar duplicados en batch
    exists = TrackingScan.query.filter_by(user_id=current_user.id, shipment_id=sh.id, source='batch').first()
    if not exists:
        record_scan(sh, user_id=current_user.id, source='batch')

    return jsonify({
        'ok': True,
        'added': (exists is None),
        'shipment': {
            'tracking_code': sh.tracking_code,
            'order_name': sh.order_name,
            'status': sh.status,
            'status_label': label_status(sh.status),
        }
    })


@bp.post('/cadeteria/api/batch_remove')
@login_required
@view_required('rastreo_cadeteria')
def cadeteria_batch_remove():
    data = request.get_json(silent=True) or {}
    code = _parse_scanned_code(data.get('code') or '')
    if not code:
        return jsonify({'ok': False, 'error': 'code_required'}), 400
    sh = _find_shipment_by_any_code(code)
    if not sh:
        return jsonify({'ok': False, 'error': 'not_found'}), 404
    TrackingScan.query.filter_by(user_id=current_user.id, shipment_id=sh.id, source='batch').delete(synchronize_session=False)
    db.session.commit()
    return jsonify({'ok': True})


@bp.get('/cadeteria/api/batch_list')
@login_required
@view_required('rastreo_cadeteria')
def cadeteria_batch_list():
    # ordenar por primer escaneo (min scanned_at)
    sub = (
        db.session.query(
            TrackingScan.shipment_id,
            db.func.min(TrackingScan.scanned_at).label('first_scan'),
        )
        .filter(TrackingScan.user_id == current_user.id, TrackingScan.source == 'batch')
        .group_by(TrackingScan.shipment_id)
        .subquery()
    )
    rows = (
        db.session.query(TrackingShipment, sub.c.first_scan)
        .join(sub, TrackingShipment.id == sub.c.shipment_id)
        .order_by(sub.c.first_scan.asc())
        .all()
    )

    items = []
    for sh, first_scan in rows:
        items.append({
            'tracking_code': sh.tracking_code,
            'order_name': sh.order_name,
            'status': sh.status,
            'status_label': label_status(sh.status),
            'first_scan': (first_scan.isoformat() if first_scan else None),
        })

    # active (si existe)
    act = TrackingScan.query.filter_by(user_id=current_user.id, source='active').order_by(TrackingScan.scanned_at.desc()).first()
    active_code = None
    if act:
        sh = TrackingShipment.query.get(act.shipment_id)
        if sh:
            active_code = sh.tracking_code

    return jsonify({'ok': True, 'items': items, 'active_code': active_code})


@bp.post('/cadeteria/api/start_reparto')
@login_required
@view_required('rastreo_cadeteria')
def cadeteria_start_reparto():
    # marca todos los del batch como OUT_FOR_DELIVERY
    ship_ids = [r.shipment_id for r in TrackingScan.query.filter_by(user_id=current_user.id, source='batch').order_by(TrackingScan.scanned_at.asc()).all()]
    if not ship_ids:
        return jsonify({'ok': False, 'error': 'empty_batch'}), 400

    # dedupe manteniendo orden
    seen = set()
    ordered_ids = []
    for sid in ship_ids:
        if sid in seen:
            continue
        seen.add(sid)
        ordered_ids.append(sid)

    ok_list = []
    err_list = []

    for sid in ordered_ids:
        sh = TrackingShipment.query.get(sid)
        if not sh:
            continue
        if sh.status in {'OUT_FOR_DELIVERY', 'ON_ROUTE_TO_DELIVERY', 'DELIVERED', 'DELIVERY_FAILED', 'RETURNED'}:
            ok_list.append({'tracking_code': sh.tracking_code, 'skipped': True, 'status': sh.status})
            continue
        try:
            add_event(sh, event_type='OUT_FOR_DELIVERY', note='Inicio de reparto (lote)', created_by_user_id=current_user.id)
            ok_list.append({'tracking_code': sh.tracking_code, 'skipped': False, 'status': sh.status})
        except Exception as ex:
            err_list.append({'tracking_code': sh.tracking_code, 'error': str(ex)})

    return jsonify({'ok': True, 'updated': ok_list, 'errors': err_list})


@bp.post('/cadeteria/api/select_active')
@login_required
@view_required('rastreo_cadeteria')
def cadeteria_select_active():
    data = request.get_json(silent=True) or {}
    code = _parse_scanned_code(data.get('code') or '')
    if not code:
        return jsonify({'ok': False, 'error': 'code_required'}), 400

    sh = _find_shipment_by_any_code(code)
    if not sh:
        return jsonify({'ok': False, 'error': 'not_found'}), 404

    # revertir activo anterior (si aplica)
    prev = TrackingScan.query.filter_by(user_id=current_user.id, source='active').order_by(TrackingScan.scanned_at.desc()).first()
    if prev and prev.shipment_id != sh.id:
        prev_sh = TrackingShipment.query.get(prev.shipment_id)
        if prev_sh and prev_sh.status == 'ON_ROUTE_TO_DELIVERY':
            try:
                add_event(prev_sh, event_type='BACK_TO_OUT_FOR_DELIVERY', note='Cambio de entrega activa', created_by_user_id=current_user.id)
            except Exception:
                pass

    # limpiar "active" actual y setear nuevo
    TrackingScan.query.filter_by(user_id=current_user.id, source='active').delete(synchronize_session=False)
    db.session.commit()
    record_scan(sh, user_id=current_user.id, source='active')

    # setear estado "en camino" si corresponde
    if sh.status == 'OUT_FOR_DELIVERY':
        try:
            add_event(sh, event_type='ON_ROUTE_TO_DELIVERY', note='En camino a entregar', created_by_user_id=current_user.id)
        except Exception as ex:
            return jsonify({'ok': False, 'error': str(ex)}), 400

    return jsonify({
        'ok': True,
        'shipment': {
            'tracking_code': sh.tracking_code,
            'order_name': sh.order_name,
            'status': sh.status,
            'status_label': label_status(sh.status),
        }
    })


@bp.post('/cadeteria/api/mark')
@login_required
@view_required('rastreo_cadeteria')
def cadeteria_api_mark():
    data = request.get_json(silent=True) or {}
    code = _parse_scanned_code(data.get('code') or '')
    event_type = (data.get('event_type') or '').strip()
    if not code:
        return jsonify({'ok': False, 'error': 'code_required'}), 400
    if event_type not in {'DELIVERED', 'DELIVERY_FAILED', 'RETURNED'}:
        return jsonify({'ok': False, 'error': 'invalid_event_type'}), 400

    sh = _find_shipment_by_any_code(code)
    if not sh:
        return jsonify({'ok': False, 'error': 'not_found'}), 404

    note = (data.get('note') or '').strip() or None
    payload = None

    if event_type == 'DELIVERY_FAILED':
        reason = (data.get('reason') or '').strip()
        detail = (data.get('detail') or '').strip() or None
        if not reason:
            return jsonify({'ok': False, 'error': 'reason_required'}), 400
        note = reason
        payload = {'reason': reason, 'detail': detail}

    if event_type == 'DELIVERED':
        receiver_relation = (data.get('receiver_relation') or '').strip() or None
        receiver_name = (data.get('receiver_name') or '').strip() or None
        receiver_id = (data.get('receiver_id') or '').strip() or None
        if not receiver_relation or not receiver_name or not receiver_id:
            return jsonify({'ok': False, 'error': 'receiver_required'}), 400
        payload = {'receiver_relation': receiver_relation, 'receiver_name': receiver_name, 'receiver_id': receiver_id}

    try:
        add_event(sh, event_type=event_type, note=note, payload=payload, created_by_user_id=current_user.id)
    except Exception as ex:
        return jsonify({'ok': False, 'error': str(ex)}), 400

    # si era el activo, limpiarlo
    TrackingScan.query.filter_by(user_id=current_user.id, source='active', shipment_id=sh.id).delete(synchronize_session=False)
    db.session.commit()

    return jsonify({
        'ok': True,
        'shipment': {
            'tracking_code': sh.tracking_code,
            'order_name': sh.order_name,
            'status': sh.status,
            'status_label': label_status(sh.status),
        }
    })
@bp.get("/cadeteria/<tracking_code>")
@login_required
@view_required("rastreo_cadeteria")
def cadeteria_detail(tracking_code: str):
    sh = _shipment_by_code(tracking_code)
    # registrar escaneo
    record_scan(sh, user_id=current_user.id, source="view")
    _configure_engine()
    order_info = {}
    try:
        order_info = etiquetas_legacy_engine._odoo_order_detail(sh.odoo_order_id)
    except Exception:
        order_info = {"pedido": sh.order_name}
    events = TrackingEvent.query.filter_by(shipment_id=sh.id).order_by(TrackingEvent.created_at.desc()).all()
    return render_template("rastreo/detail.html", title="Rastreo - Cadetería", mode="cadeteria", shipment=sh, order=order_info, events=events)


@bp.post("/cadeteria/<tracking_code>/event")
@login_required
@view_required("rastreo_cadeteria")
def cadeteria_event(tracking_code: str):
    sh = _shipment_by_code(tracking_code)
    event_type = (request.form.get("event_type") or "").strip()
    note = (request.form.get("note") or "").strip() or None

    if event_type not in {"OUT_FOR_DELIVERY", "DELIVERED", "DELIVERY_FAILED", "RETURNED"}:
        abort(400)

    payload = None
    image_path = None
    if event_type == "DELIVERY_FAILED":
        f = request.files.get("photo")
        if f and f.filename:
            up_dir = Path(current_app.root_path).parent / current_app.config.get("GENERATED_DIR", "generated") / "tracking_photos"
            up_dir.mkdir(parents=True, exist_ok=True)
            ext = os.path.splitext(f.filename)[1].lower() or ".jpg"
            fname = f"fail_{tracking_code}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}{ext}"
            fp = up_dir / fname
            f.save(fp)
            image_path = str(fp)

    if event_type == "DELIVERED":
        receiver_relation = (request.form.get("receiver_relation") or "").strip() or None
        receiver_name = (request.form.get("receiver_name") or "").strip() or None
        receiver_id = (request.form.get("receiver_id") or "").strip() or None
        if not receiver_relation or not receiver_name or not receiver_id:
            abort(400, "receiver_required")
        payload = {"receiver_relation": receiver_relation, "receiver_name": receiver_name, "receiver_id": receiver_id}

    try:
        add_event(sh, event_type=event_type, note=note, payload=payload, created_by_user_id=current_user.id, image_path=image_path)
    except ValueError as ex:
        abort(400, str(ex))
    return redirect(url_for("rastreo.cadeteria_detail", tracking_code=tracking_code))


@bp.get("/cadeteria/mis-escaneos")
@login_required
@view_required("rastreo_cadeteria")
def my_scans():
    day = (request.args.get("day") or "").strip()
    tz = get_app_timezone()
    try:
        local_day = date.fromisoformat(day) if day else datetime.now(tz).date()
    except Exception:
        local_day = datetime.now(tz).date()

    # scanned_at se guarda como naive UTC (datetime.utcnow). Convertimos el rango local -> UTC naive.
    start_local = datetime.combine(local_day, time.min).replace(tzinfo=tz)
    end_local = datetime.combine(local_day, time.max).replace(tzinfo=tz)
    d0 = start_local.astimezone(timezone.utc).replace(tzinfo=None)
    d1 = end_local.astimezone(timezone.utc).replace(tzinfo=None)

    scans = TrackingScan.query.filter(
        TrackingScan.user_id == current_user.id,
        TrackingScan.scanned_at >= d0,
        TrackingScan.scanned_at <= d1,
    ).order_by(TrackingScan.scanned_at.desc()).limit(300).all()

    # map shipments
    ship_ids = [s.shipment_id for s in scans]
    shipments = {sh.id: sh for sh in TrackingShipment.query.filter(TrackingShipment.id.in_(ship_ids)).all()} if ship_ids else {}

    rows = []
    for s in scans:
        sh = shipments.get(s.shipment_id)
        if not sh:
            continue
        rows.append({
            "scanned_at": s.scanned_at,
            "tracking_code": sh.tracking_code,
            "order_name": sh.order_name,
            "status": sh.status,
        })

    return render_template("rastreo/my_scans.html", title="Mis escaneos", day=local_day.isoformat(), rows=rows)


@bp.get("/ventas")
@login_required
@view_required("rastreo_ventas")
def ventas():
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip()

    from sqlalchemy.orm import joinedload

    # Paginación (para histórico completo)
    #
    # Para que los filtros de la grilla (por columna) se apliquen sobre
    # todo el conjunto de resultados y no sólo sobre la página visible,
    # se permite incrementar el `per_page` a un valor elevado.  Si el
    # usuario no especifica `per_page` se usa un valor por defecto
    # suficientemente grande (50 000).  Esto hace que la consulta
    # devuelva prácticamente todos los envíos disponibles, de modo que
    # los filtros en el frontend afecten a todos los registros cargados.
    try:
        page = int(request.args.get("page") or 1)
    except Exception:
        page = 1
    page = max(page, 1)
    try:
        per_page = int(request.args.get("per_page") or 50000)
    except Exception:
        per_page = 50000
    # Limitar per_page entre 50 y 100 000 para evitar abusos.
    per_page = max(50, min(per_page, 100000))

    query = TrackingShipment.query.options(
        joinedload(TrackingShipment.warehouse_user),
        joinedload(TrackingShipment.courier_user),
    )
    if q:
        query = query.filter(
            (TrackingShipment.tracking_code.ilike(f"%{q}%")) |
            (TrackingShipment.order_name.ilike(f"%{q}%")) |
            (TrackingShipment.id_web.ilike(f"%{q}%"))
        )
    if status:
        query = query.filter(TrackingShipment.status == status)

    # Recuperar filas.  Si `per_page` es muy grande, la consulta
    # devolverá todos los registros (sin paginación real) para que los
    # filtros de la tabla se apliquen al conjunto completo.
    rows = (
        query
        .order_by(TrackingShipment.updated_at.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    # Total (para habilitar "Siguiente").  Evitar error cuando la
    # consulta es compleja.
    try:
        total = query.count()
    except Exception:
        total = None

    delivered_by_map = {}
    delivered_dt_map = {}  # shipment_id -> datetime (created_at del evento DELIVERED)
    try:
        ship_ids = [s.id for s in rows]
        if ship_ids:
            from sqlalchemy.orm import joinedload
            sub = (
                db.session.query(
                    TrackingEvent.shipment_id,
                    db.func.max(TrackingEvent.id).label("max_id"),
                )
                .filter(TrackingEvent.shipment_id.in_(ship_ids), TrackingEvent.event_type == "DELIVERED")
                .group_by(TrackingEvent.shipment_id)
                .subquery()
            )
            latest = (
                db.session.query(TrackingEvent)
                .join(sub, db.and_(TrackingEvent.shipment_id == sub.c.shipment_id, TrackingEvent.id == sub.c.max_id))
                .options(joinedload(TrackingEvent.created_by))
                .all()
            )
            for ev in latest:
                try:
                    payload = json.loads(ev.payload_json) if ev.payload_json else {}
                except Exception:
                    payload = {}
                name = payload.get("delivered_courier_name") or (ev.created_by.full_name if ev.created_by else None)
                if name:
                    delivered_by_map[int(ev.shipment_id)] = name
                delivered_dt_map[int(ev.shipment_id)] = ev.created_at
    except Exception:
        delivered_by_map = {}

    # Cliente / Dirección: intentar resolver el nombre y la dirección de
    # entrega para cada envío.  En versiones anteriores sólo se
    # consultaba por BatchOrder.sale_order_id, lo que dejaba muchos
    # pedidos sin información cuando no habían sido importados como
    # batch.  Ahora también se consulta por id_web y sale_order_ref,
    # y se indexa por shipment.id para facilitar el acceso en la
    # plantilla.
    cliente_map: dict[int, str] = {}
    direccion_map: dict[int, str] = {}
    try:
        # Recopilar claves para búsqueda en batch_orders
        odoo_ids = [int(s.odoo_order_id) for s in rows if getattr(s, "odoo_order_id", None)]
        id_webs = [s.id_web for s in rows if getattr(s, "id_web", None)]
        order_refs = [s.order_name for s in rows if getattr(s, "order_name", None)]
        # Construir la consulta con OR para abarcar todos los casos
        conds = []
        if odoo_ids:
            conds.append(BatchOrder.sale_order_id.in_(odoo_ids))
        if id_webs:
            conds.append(BatchOrder.id_web.in_(id_webs))
        if order_refs:
            conds.append(BatchOrder.sale_order_ref.in_(order_refs))
        bos = []
        if conds:
            bos = BatchOrder.query.filter(db.or_(*conds)).all()
        # Construir índices por tipo de clave
        by_sale_id: dict[int, BatchOrder] = {}
        by_id_web: dict[str, BatchOrder] = {}
        by_ref: dict[str, BatchOrder] = {}
        for bo in bos:
            try:
                if bo.sale_order_id is not None:
                    by_sale_id[int(bo.sale_order_id)] = bo
            except Exception:
                pass
            if bo.id_web:
                by_id_web[str(bo.id_web)] = bo
            if bo.sale_order_ref:
                by_ref[str(bo.sale_order_ref)] = bo
        # Asignar cliente/dirección a cada shipment usando la mejor coincidencia
        for sh in rows:
            bo: BatchOrder | None = None
            # Prioridad: sale_order_id
            try:
                sid = int(sh.odoo_order_id)
                bo = by_sale_id.get(sid)
            except Exception:
                bo = None
            # Segundo: id_web
            if bo is None and sh.id_web:
                bo = by_id_web.get(str(sh.id_web))
            # Tercero: order_name/sale_order_ref
            if bo is None and sh.order_name:
                bo = by_ref.get(str(sh.order_name))
            if bo:
                if bo.cliente:
                    cliente_map[sh.id] = bo.cliente
                if bo.direccion:
                    direccion_map[sh.id] = bo.direccion
    except Exception:
        # Ante cualquier error, dejar los mapas vacíos para evitar rupturas
        cliente_map = {}
        direccion_map = {}

    # URLs de paginación preservando filtros
    def _build_page_url(new_page: int):
        args = request.args.to_dict(flat=True)
        args["page"] = str(new_page)
        args["per_page"] = str(per_page)
        return url_for("rastreo.ventas", **args)

    prev_url = _build_page_url(page - 1) if page > 1 else None
    has_next = (total is None) or ((page * per_page) < total)
    next_url = _build_page_url(page + 1) if has_next else None
    base_url = (os.environ.get("PUBLIC_TRACKING_BASE_URL") or "https://rastreo.iluminaras.cloud").strip()
    return render_template(
        "rastreo/ventas.html",
        title="Administración de Tracking",
        rows=rows,
        delivered_by_map=delivered_by_map,
        delivered_dt_map=delivered_dt_map,
        cliente_map=cliente_map,
        direccion_map=direccion_map,
        q=q,
        status=status,
        public_tracking_base_url=base_url,
        page=page,
        per_page=per_page,
        total=total,
        prev_url=prev_url,
        next_url=next_url,
    )


@bp.get("/ventas/partial")
@login_required
@view_required("rastreo_ventas")
def ventas_partial():
    """Fragmento HTML para auto-refresh de la grilla de Ventas."""
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip()
    from sqlalchemy.orm import joinedload

    # Respeta paginación actual (misma querystring)
    try:
        page = int(request.args.get("page") or 1)
    except Exception:
        page = 1
    page = max(page, 1)
    try:
        per_page = int(request.args.get("per_page") or 50000)
    except Exception:
        per_page = 50000
    # Limitar per_page para evitar abusos.  Por defecto se entregan
    # prácticamente todos los envíos para que el filtrado en el
    # frontend se aplique sobre el conjunto completo.
    per_page = max(50, min(per_page, 100000))
    query = TrackingShipment.query.options(
        joinedload(TrackingShipment.warehouse_user),
        joinedload(TrackingShipment.courier_user),
    )
    if q:
        query = query.filter(
            (TrackingShipment.tracking_code.ilike(f"%{q}%")) |
            (TrackingShipment.order_name.ilike(f"%{q}%")) |
            (TrackingShipment.id_web.ilike(f"%{q}%"))
        )
    if status:
        query = query.filter(TrackingShipment.status == status)
    rows = (
        query
        .order_by(TrackingShipment.updated_at.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )
    delivered_by_map = {}
    delivered_dt_map = {}
    try:
        ship_ids = [s.id for s in rows]
        if ship_ids:
            from sqlalchemy.orm import joinedload
            sub = (
                db.session.query(
                    TrackingEvent.shipment_id,
                    db.func.max(TrackingEvent.id).label("max_id"),
                )
                .filter(TrackingEvent.shipment_id.in_(ship_ids), TrackingEvent.event_type == "DELIVERED")
                .group_by(TrackingEvent.shipment_id)
                .subquery()
            )
            latest = (
                db.session.query(TrackingEvent)
                .join(sub, db.and_(TrackingEvent.shipment_id == sub.c.shipment_id, TrackingEvent.id == sub.c.max_id))
                .options(joinedload(TrackingEvent.created_by))
                .all()
            )
            for ev in latest:
                try:
                    payload = json.loads(ev.payload_json) if ev.payload_json else {}
                except Exception:
                    payload = {}
                name = payload.get("delivered_courier_name") or (ev.created_by.full_name if ev.created_by else None)
                if name:
                    delivered_by_map[int(ev.shipment_id)] = name
                delivered_dt_map[int(ev.shipment_id)] = ev.created_at
    except Exception:
        delivered_by_map = {}

    # Cliente / Dirección: resolver nombre y dirección de entrega por envío.
    cliente_map: dict[int, str] = {}
    direccion_map: dict[int, str] = {}
    try:
        odoo_ids = [int(s.odoo_order_id) for s in rows if getattr(s, "odoo_order_id", None)]
        id_webs = [s.id_web for s in rows if getattr(s, "id_web", None)]
        order_refs = [s.order_name for s in rows if getattr(s, "order_name", None)]
        conds = []
        if odoo_ids:
            conds.append(BatchOrder.sale_order_id.in_(odoo_ids))
        if id_webs:
            conds.append(BatchOrder.id_web.in_(id_webs))
        if order_refs:
            conds.append(BatchOrder.sale_order_ref.in_(order_refs))
        bos = []
        if conds:
            bos = BatchOrder.query.filter(db.or_(*conds)).all()
        by_sale_id: dict[int, BatchOrder] = {}
        by_id_web: dict[str, BatchOrder] = {}
        by_ref: dict[str, BatchOrder] = {}
        for bo in bos:
            try:
                if bo.sale_order_id is not None:
                    by_sale_id[int(bo.sale_order_id)] = bo
            except Exception:
                pass
            if bo.id_web:
                by_id_web[str(bo.id_web)] = bo
            if bo.sale_order_ref:
                by_ref[str(bo.sale_order_ref)] = bo
        for sh in rows:
            bo = None
            try:
                sid = int(sh.odoo_order_id)
                bo = by_sale_id.get(sid)
            except Exception:
                bo = None
            if bo is None and sh.id_web:
                bo = by_id_web.get(str(sh.id_web))
            if bo is None and sh.order_name:
                bo = by_ref.get(str(sh.order_name))
            if bo:
                if bo.cliente:
                    cliente_map[sh.id] = bo.cliente
                if bo.direccion:
                    direccion_map[sh.id] = bo.direccion
    except Exception:
        cliente_map = {}
        direccion_map = {}

    return render_template(
        "rastreo/_ventas_rows.html",
        rows=rows,
        delivered_by_map=delivered_by_map,
        delivered_dt_map=delivered_dt_map,
        cliente_map=cliente_map,
        direccion_map=direccion_map,
    )


@bp.post("/ventas/export_xlsx")
@login_required
@view_required("rastreo_ventas")
def ventas_export_xlsx():
    """Exporta a XLSX los registros actualmente filtrados/visibles en la grilla (desde el frontend)."""
    payload = request.get_json(silent=True) or {}
    if not isinstance(payload, dict):
        payload = {}

    rows = payload.get("rows") or []
    if not isinstance(rows, list):
        rows = []

    # hard limit para evitar abusos
    if len(rows) > 20000:
        return jsonify({"error": "Demasiadas filas para exportar (máx 20000)"}), 400
    if not rows:
        return jsonify({"error": "No hay filas para exportar"}), 404

    from io import BytesIO
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    ws = wb.active
    ws.title = "Tracking"

    headers = [
        "Fecha",
        "Hora",
        "Código",
        "Pedido",
        "Cliente",
        "Dirección entrega",
        "Fecha entrega",
        "Hora entrega",
        "Armado por",
        "Cadete",
        "Estado",
        "Link público",
    ]
    ws.append(headers)

    def _safe_str(v):
        if v is None:
            return ""
        return str(v)

    for r in rows:
        if not isinstance(r, dict):
            continue
        ws.append(
            [
                _safe_str(r.get("fecha")),
                _safe_str(r.get("hora")),
                _safe_str(r.get("codigo")),
                _safe_str(r.get("pedido")),
                _safe_str(r.get("cliente")),
                _safe_str(r.get("direccion_entrega")),
                _safe_str(r.get("fecha_entrega")),
                _safe_str(r.get("hora_entrega")),
                _safe_str(r.get("armado_por")),
                _safe_str(r.get("cadete")),
                _safe_str(r.get("estado")),
                _safe_str(r.get("link_publico")),
            ]
        )

    # Tabla estilo Excel
    last_row = ws.max_row
    last_col = len(headers)
    ref = f"A1:{get_column_letter(last_col)}{last_row}"
    tab = Table(displayName="TrackingTable", ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    # Autosize simple
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 80)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"tracking_ventas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.get("/ventas/<tracking_code>")
@login_required
@view_required("rastreo_ventas")
def ventas_detail(tracking_code: str):
    sh = _shipment_by_code(tracking_code)
    _configure_engine()
    order_info = {}
    try:
        order_info = etiquetas_legacy_engine._odoo_order_detail(sh.odoo_order_id)
    except Exception:
        order_info = {"pedido": sh.order_name}
    events = TrackingEvent.query.filter_by(shipment_id=sh.id).order_by(TrackingEvent.created_at.desc()).all()
    override_statuses = [
        "LABEL_CREATED",
        "PICKING_STARTED",
        "STOCK_MISSING",
        "STOCK_RESOLVED",
        "READY_FOR_DISPATCH",
        "OUT_FOR_DELIVERY",
        "ON_ROUTE_TO_DELIVERY",
        "DELIVERY_FAILED",
        "RETURNED",
        "DELIVERED",
    ]
    return render_template(
        "rastreo/detail.html",
        title="Administración de Tracking",
        mode="ventas",
        shipment=sh,
        order=order_info,
        events=events,
        override_statuses=override_statuses,
    )


@bp.post("/ventas/<tracking_code>/decision")
@login_required
@view_required("rastreo_ventas")
def ventas_decision(tracking_code: str):
    sh = _shipment_by_code(tracking_code)
    decision = (request.form.get("decision") or "").strip()
    note = (request.form.get("note") or "").strip() or None
    if decision not in {"CONTINUE", "HOLD", "CANCEL"}:
        abort(400)
    payload = {"decision": decision}
    try:
        add_event(sh, event_type="SALES_DECISION", note=note, payload=payload, created_by_user_id=current_user.id)
    except ValueError as ex:
        abort(400, str(ex))
    return redirect(url_for("rastreo.ventas_detail", tracking_code=tracking_code))


@bp.post("/ventas/<tracking_code>/override")
@login_required
@view_required("rastreo_ventas")
def ventas_override(tracking_code: str):
    sh = _shipment_by_code(tracking_code)
    new_status = (request.form.get("new_status") or "").strip()
    note = (request.form.get("note") or "").strip() or None
    try:
        ventas_override_status(sh, new_status=new_status, note=note, created_by_user_id=current_user.id)
    except ValueError as ex:
        abort(400, str(ex))
    return redirect(url_for("rastreo.ventas_detail", tracking_code=tracking_code))


@bp.post("/ventas/<tracking_code>/reset")
@login_required
@view_required("rastreo_ventas")
def ventas_reset(tracking_code: str):
    sh = _shipment_by_code(tracking_code)
    note = (request.form.get("note") or "").strip() or None
    try:
        ventas_reset_to_zero(sh, note=note, created_by_user_id=current_user.id)
    except ValueError as ex:
        abort(400, str(ex))
    return redirect(url_for("rastreo.ventas_detail", tracking_code=tracking_code))


@bp.get("/pedido/<tracking_code>")
@login_required
def pedido_view(tracking_code: str):
    sh = _shipment_by_code(tracking_code)
    data = get_order_full(sh.odoo_order_id)
    return render_template("rastreo/pedido.html", title=f"Pedido {data.get('name') or sh.order_name}", order=data, shipment=sh)


@bp.get("/photo")
@login_required
def show_photo():
    # servir fotos guardadas (solo a usuarios logueados)
    p = (request.args.get("p") or "").strip()
    if not p:
        abort(404)
    base = Path(current_app.root_path).parent / current_app.config.get("GENERATED_DIR", "generated") / "tracking_photos"
    try:
        fp = Path(p)
    except Exception:
        abort(404)
    if not fp.is_absolute():
        fp = base / p
    fp = fp.resolve()
    if base.resolve() not in fp.parents:
        abort(403)
    if not fp.exists():
        abort(404)
    return send_file(str(fp), as_attachment=False)


@bp.post("/api/decode_qr")
@login_required
def api_decode_qr():
    f = request.files.get("image")
    if not f:
        return jsonify({"ok": False, "error": "image_required"}), 400

    b = f.read()
    if not b:
        return jsonify({"ok": False, "error": "empty"}), 400

    try:
        import numpy as np  # optional dependency
        import cv2  # optional dependency
    except Exception:
        return jsonify({"ok": False, "error": "qr_decode_deps_missing"}), 501

    try:
        arr = np.frombuffer(b, dtype=np.uint8)
        img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    except Exception:
        img = None

    code = None

    # OpenCV QRCodeDetector
    try:
        if not code and img is not None:
            det = cv2.QRCodeDetector()
            val, _pts, _ = det.detectAndDecode(img)
            if val:
                code = val.strip()
    except Exception:
        pass

    if not code:
        return jsonify({"ok": False, "error": "not_found"}), 404

    # si el QR contiene URL /rastreo/go/<code>, extraer último segmento
    if "/rastreo/go/" in code:
        try:
            code = code.split("/rastreo/go/")[-1].split("?")[0].strip("/")
        except Exception:
            pass

    return jsonify({"ok": True, "code": code})
